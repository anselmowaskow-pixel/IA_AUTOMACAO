from openpyxl import Workbook
from pathlib import Path

PLANILHA = Path(r"C:\IA_AUTOMACAO\STORAGE\Empresa A\RESULTADOS\EMPRESA_CONTROLE_GERAL.xlsx")

wb = Workbook()

# Aba EVENTOS
ws_eventos = wb.active
ws_eventos.title = "EVENTOS"
ws_eventos.append([
    "id_evento",
    "arquivo",
    "extensao",
    "hash",
    "data_entrada",
    "processo",
    "origem",
    "destino"
])

# Aba ENTRADA
ws_entrada = wb.create_sheet("ENTRADA")
ws_entrada.append([
    "arquivo",
    "data",
    "status"
])

# Aba SAIDA
ws_saida = wb.create_sheet("SAIDA")
ws_saida.append([
    "arquivo",
    "data",
    "status"
])

wb.save(PLANILHA)

print("Planilha-mãe criada com sucesso:", PLANILHA)