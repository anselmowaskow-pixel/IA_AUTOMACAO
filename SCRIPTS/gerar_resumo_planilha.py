from openpyxl import load_workbook
from pathlib import Path
from collections import Counter
from datetime import datetime

PLANILHA = Path(r"C:\IA_AUTOMACAO\STORAGE\Empresa A\RESULTADOS\EMPRESA_CONTROLE_GERAL.xlsx")

def gerar_resumo():
    wb = load_workbook(PLANILHA)

    if "EVENTOS" not in wb.sheetnames:
        print("ERRO: Aba EVENTOS não encontrada")
        return

    ws_eventos = wb["EVENTOS"]

    if "RESUMO" in wb.sheetnames:
        ws_resumo = wb["RESUMO"]
        ws_resumo.delete_rows(1, ws_resumo.max_row)
    else:
        ws_resumo = wb.create_sheet("RESUMO")

    extensoes = []
    datas = []

    for row in ws_eventos.iter_rows(min_row=2, values_only=True):
        extensao = row[2]   # coluna extensao
        data_evt = row[4]  # coluna data_entrada

        if extensao:
            extensoes.append(extensao)

        if data_evt:
            datas.append(data_evt)

    contagem = Counter(extensoes)

    ws_resumo.append(["RESUMO GERAL"])
    ws_resumo.append(["Gerado em", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_resumo.append([])
    ws_resumo.append(["Total de arquivos", len(extensoes)])
    ws_resumo.append([])

    ws_resumo.append(["Por tipo"])
    ws_resumo.append(["Extensão", "Quantidade"])

    for ext, qtd in contagem.items():
        ws_resumo.append([ext, qtd])

    if datas:
        ws_resumo.append([])
        ws_resumo.append(["Último processamento", max(datas)])

    wb.save(PLANILHA)
    print("RESUMO atualizado com sucesso.")

if __name__ == "__main__":
    gerar_resumo()